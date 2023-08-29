from collections import Counter
from openpyxl import load_workbook
import string

def remove_punctuation(word):
    return word.translate(str.maketrans('', '', string.punctuation)).replace("–","")

TOP_WORDS = 1000

# Load the workbook
workbook = load_workbook('input.xlsx')

words_by_sheet = []
# Iterate over all sheets
for sheet in workbook:
    # I only want to analyze tabs in my sheet that I named as a number,
    # so added isdigit()
    if (sheet.title.isdigit()):
        words = []
        # Iterate over all rows in the sheet
        for row in sheet.iter_rows():
            # Iterate over all cells in the row
            for cell in row:
                if cell.value is not None:
                    for word in str(cell.value).split():
                        fixedWord = remove_punctuation(word).lower()
                        if len(fixedWord) > 0:
                            words.extend([fixedWord])
        words_by_sheet.append(words)

# Count the number of sheets each word appears in
sheet_counts = Counter()
for words in words_by_sheet:
    sheet_words = set(words)
    for word in sheet_words:
            sheet_counts[word] += 1

# Print the top words by number of sheets they appear in
for word, count in sheet_counts.most_common(TOP_WORDS):
    print(f'{word}: {count}')
